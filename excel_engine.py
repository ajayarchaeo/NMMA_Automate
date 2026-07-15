import os
import pandas as pd
import openpyxl
from copy import copy
from PySide6.QtCore import QThread, Signal
from template_mapper import TemplateMapper


def audit_record(row_dict, mapping_dict):
    """
    Checks if crucial fields for an antiquity record are populated.
    Returns a list of missing field names.
    """
    crucial_fields = {
        "C3": "Title",
        "C4": "Type of Object",
        "C8": "Material",
        "C10": "Description",
        "C16": "Accession No",
        "C5": "Date/Period"
    }
    missing_fields = []
    
    def value(column_name):
        v = row_dict.get(column_name, "")
        if str(v) == "nan" or v is None:
            return ""
        return str(v).strip()

    for cell_ref, display_name in crucial_fields.items():
        col_name = mapping_dict.get(cell_ref)
        val = ""
        if col_name and col_name != "(Use Template Default)":
            val = value(col_name)
            
        # Accession C16 fallback check
        if cell_ref == "C16" and not val:
            val = value("Field Accession No.")
            if not val:
                val = value("NMMA No.")
                
        if not val:
            missing_fields.append(display_name)
            
    return missing_fields


def find_photo_files(record, index, photos_dir, mapping_dict):
    """
    Looks inside photos_dir for all image files whose name matches the antiquity record:
    1. The filename specified in the mapped "Photograph" column.
    2. The resolved Accession/Registration No, Field Accession No., or NMMA No.
    3. The 1-based index (e.g., "00001" or "1").
    Matches exact filenames or filenames with suffixes (e.g. "35a", "35b", "35_1").
    Returns a sorted list of absolute paths to all matching images, or empty list.
    """
    if not photos_dir or not os.path.exists(photos_dir):
        return []
        
    def value(column_name):
        if column_name in record:
            v = record[column_name]
            if str(v) == "nan" or v is None:
                return ""
            return str(v).strip()
        for col in record.index:
            if str(col).lower() == str(column_name).lower():
                v = record[col]
                if str(v) == "nan" or v is None:
                    return ""
                return str(v).strip()
        return ""

    candidates = []
    
    # 1. Mapped photograph column value
    photo_col = mapping_dict.get("C13")
    if photo_col and photo_col != "(Use Template Default)":
        photo_val = value(photo_col)
        if photo_val:
            candidates.append(photo_val)
            
    # 2. Accession values
    acc_col = mapping_dict.get("C16")
    acc_val = ""
    if acc_col and acc_col != "(Use Template Default)":
        acc_val = value(acc_col)
    if not acc_val:
        acc_val = value("Field Accession No.")
        if not acc_val:
            acc_val = value("NMMA No.")
    if acc_val:
        candidates.append(acc_val)
        
    # 3. Row index
    candidates.append(f"{index:05d}")
    candidates.append(str(index))

    # Supported photo extensions
    valid_exts = {".jpg", ".jpeg", ".png", ".gif", ".bmp"}
    
    try:
        files = os.listdir(photos_dir)
    except Exception:
        return []

    import re
    matched_paths = []
    
    for cand in candidates:
        cand_clean = cand.strip().lower()
        if not cand_clean:
            continue
            
        _, ext = os.path.splitext(cand_clean)
        if ext in valid_exts:
            for f in files:
                if f.lower() == cand_clean:
                    matched_paths.append(os.path.join(photos_dir, f))
            if matched_paths:
                break
                
        pattern = re.compile(rf"^{re.escape(cand_clean)}(?:[a-zA-Z]|_\d+|-\d+|_\w+|-\w+)?$")
        for f in files:
            f_name, f_ext = os.path.splitext(f.lower())
            if f_ext in valid_exts and pattern.match(f_name):
                matched_paths.append(os.path.join(photos_dir, f))
                
        if matched_paths:
            break
            
    return sorted(list(set(matched_paths)))


def get_record_photo(record, index, photos_dir, mapping_dict, output_folder):
    """
    Finds and resolves the photograph path for a record.
    If multiple photographs match, merges them side-by-side using Pillow
    and returns the temporary merged image path.
    """
    image_paths = find_photo_files(record, index, photos_dir, mapping_dict)
    if not image_paths:
        return None
        
    if len(image_paths) == 1:
        return image_paths[0]
        
    from PIL import Image as PILImage
    try:
        images = []
        for path in image_paths:
            try:
                img = PILImage.open(path)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                images.append(img)
            except Exception as e:
                print(f"Error opening image {path}: {e}")
                
        if not images:
            return None
            
        if len(images) == 1:
            return image_paths[0]
            
        canvas_w = 600
        canvas_h = 450
        combined_img = PILImage.new('RGB', (canvas_w, canvas_h), color='#ffffff')
        
        num_images = len(images)
        spacing = 15
        total_spacings_w = spacing * (num_images - 1)
        slot_w = int((canvas_w - total_spacings_w) / num_images)
        
        for i, img in enumerate(images):
            orig_w, orig_h = img.size
            scale = min(slot_w / orig_w, canvas_h / orig_h)
            
            new_w = int(orig_w * scale)
            new_h = int(orig_h * scale)
            
            resized_img = img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
            
            slot_x_start = i * (slot_w + spacing)
            x_offset = slot_x_start + int((slot_w - new_w) / 2)
            y_offset = int((canvas_h - new_h) / 2)
            
            combined_img.paste(resized_img, (x_offset, y_offset))
            
        temp_dir = os.path.join(output_folder, ".temp_photos")
        os.makedirs(temp_dir, exist_ok=True)
        temp_filename = f"merged_{index}_{num_images}.jpg"
        temp_path = os.path.join(temp_dir, temp_filename)
        combined_img.save(temp_path, "JPEG", quality=90)
        
        for img in images:
            img.close()
            
        return temp_path
    except Exception as e:
        print(f"Failed to merge images for index {index}: {e}")
        return image_paths[0] if image_paths else None


def copy_block(ws, src_start, src_end, dest_start):
    """
    Helper function to copy a range of rows vertically (including values, formatting,
    fonts, borders, fills, alignments, and row heights).
    """
    row_offset = dest_start - src_start
    for r in range(src_start, src_end + 1):
        dest_row = r + row_offset
        
        if r in ws.row_dimensions:
            ws.row_dimensions[dest_row].height = ws.row_dimensions[r].height
            
        for c in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=r, column=c)
            dest_cell = ws.cell(row=dest_row, column=c)
            
            dest_cell.value = src_cell.value
            
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.number_format = copy(src_cell.number_format)


def copy_block_horizontal(ws, src_start_col, src_end_col, dest_start_col, max_row=22):
    """
    Helper function to copy a range of columns horizontally (including values, formatting,
    fonts, borders, fills, alignments, number formats, and column widths) up to max_row.
    """
    from openpyxl.utils import get_column_letter
    col_offset = dest_start_col - src_start_col
    for c in range(src_start_col, src_end_col + 1):
        dest_col = c + col_offset
        src_letter = get_column_letter(c)
        dest_letter = get_column_letter(dest_col)
        
        # Copy column width if defined
        if src_letter in ws.column_dimensions:
            ws.column_dimensions[dest_letter].width = ws.column_dimensions[src_letter].width
            
        for r in range(1, max_row + 1):
            src_cell = ws.cell(row=r, column=c)
            dest_cell = ws.cell(row=r, column=dest_col)
            
            dest_cell.value = src_cell.value
            
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.number_format = copy(src_cell.number_format)


def process_single_row(args):
    """
    Worker function to process a single antiquity row.
    Runs in a child process.
    args: tuple of (row_dict, template_path, output_folder, index, mapping_dict, photos_dir, photo_width, photo_height)
    """
    row_dict, template_path, output_folder, index, mapping_dict, photos_dir, photo_width, photo_height = args
    try:
        wb = openpyxl.load_workbook(template_path)
        
        # Audit crucial values
        missing_fields = audit_record(row_dict, mapping_dict)
        
        # Find and resolve photographs
        record = pd.Series(row_dict)
        photo_files = find_photo_files(record, index, photos_dir, mapping_dict)
        photo_path = get_record_photo(record, index, photos_dir, mapping_dict, output_folder)
        
        if photo_path:
            if len(photo_files) > 1:
                photo_status_msg = f"Photo matched & merged ({len(photo_files)} files: {', '.join([os.path.basename(p) for p in photo_files])})"
            else:
                photo_status_msg = f"Photo matched: {os.path.basename(photo_path)}"
        else:
            photo_status_msg = "No matching photo found"
        
        # Instantiate mapper and fill record
        mapper = TemplateMapper(wb)
        mapper.fill_record(record, mapping_dict, photo_path=photo_path, photo_width=photo_width, photo_height=photo_height)
        
        acc_col = mapping_dict.get("C16")
        accession_no = ""
        if acc_col and acc_col != "(Use Template Default)":
            accession_no = str(row_dict.get(acc_col, ""))
            
        if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
            accession_no = str(row_dict.get("Field Accession No.", ""))
        if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
            accession_no = str(row_dict.get("NMMA No.", ""))
            
        title_col = mapping_dict.get("C3")
        title = ""
        if title_col and title_col != "(Use Template Default)":
            title = str(row_dict.get(title_col, ""))
        
        def sanitize(val):
            val = str(val).strip()
            if not val or val.lower() == "nan":
                return ""
            for char in r'\/:*?"<>|':
                val = val.replace(char, "_")
            return val

        san_acc = sanitize(accession_no)
        san_title = sanitize(title)
        
        parts = [f"{index:05d}"]
        if san_acc:
            parts.append(san_acc)
        if san_title:
            parts.append(san_title[:30])
            
        filename = "_".join(parts) + ".xlsx"
        output_file = os.path.join(output_folder, filename)
        
        save_status_msg = "Success"
        try:
            wb.save(output_file)
        except PermissionError:
            import time
            base, ext = os.path.splitext(output_file)
            fallback_file = f"{base}_LOCKED_{int(time.time())}{ext}"
            wb.save(fallback_file)
            output_file = fallback_file
            save_status_msg = "Saved as copy (original open in Excel)"
            
        wb.close()
        
        return {
            "index": index,
            "success": True,
            "output_file": output_file,
            "title": title if title and title.lower() != "nan" else "Unknown Title",
            "resolved_acc": accession_no if accession_no and accession_no.lower() != "nan" else "None",
            "missing_crucial": missing_fields,
            "photo_status": photo_status_msg,
            "save_status": save_status_msg,
            "error": None
        }
    except Exception as e:
        import traceback
        return {
            "index": index,
            "success": False,
            "output_file": None,
            "title": str(row_dict.get("Title of the object", "Unknown Title")),
            "resolved_acc": str(row_dict.get("NMMA No.", "None")),
            "missing_crucial": [],
            "photo_status": "Failed to process photo",
            "save_status": "Failed to save file",
            "error": f"{str(e)}\n{traceback.format_exc()}"
        }


def process_chunk(args):
    """
    Worker function to process a chunk of antiquity rows into a single workbook.
    Supports either multi-sheet tabs, vertically stacked, or horizontally stacked layout.
    Runs in a child process.
    args: tuple of (rows_list, template_path, output_folder, chunk_index, start_idx, end_idx, group_layout, mapping_dict, photos_dir, photo_width, photo_height)
    """
    rows_list, template_path, output_folder, chunk_index, start_idx, end_idx, group_layout, mapping_dict, photos_dir, photo_width, photo_height = args
    try:
        wb = openpyxl.load_workbook(template_path)
        rows_status = []
        
        # ---------------------------------------------
        # Layout: Vertically Stacked in a Single Sheet
        # ---------------------------------------------
        if group_layout == "stacked":
            ws = wb.active
            ws.title = "NMMA Forms"
            
            if ws.views.sheetView:
                ws.views.sheetView[0].showGridLines = True
            
            src_start, src_end = 1, 22
            block_height = src_end - src_start + 1
            spacing = 2
            
            for i, row_dict in enumerate(rows_list):
                row_offset = i * (block_height + spacing)
                row_idx = start_idx + i
                
                if i > 0:
                    copy_block(ws, src_start, src_end, src_start + row_offset)
                
                missing_fields = audit_record(row_dict, mapping_dict)
                record = pd.Series(row_dict)
                photo_files = find_photo_files(record, row_idx, photos_dir, mapping_dict)
                photo_path = get_record_photo(record, row_idx, photos_dir, mapping_dict, output_folder)
                
                if photo_path:
                    if len(photo_files) > 1:
                        photo_status_msg = f"Photo matched & merged ({len(photo_files)} files: {', '.join([os.path.basename(p) for p in photo_files])})"
                    else:
                        photo_status_msg = f"Photo matched: {os.path.basename(photo_path)}"
                else:
                    photo_status_msg = "No matching photo found"
                
                mapper = TemplateMapper(wb)
                mapper.sheet = ws
                mapper.fill_record(record, mapping_dict, row_offset=row_offset, photo_path=photo_path, photo_width=photo_width, photo_height=photo_height)
                
                if i < len(rows_list) - 1:
                    from openpyxl.worksheet.pagebreak import Break
                    ws.row_breaks.append(Break(id=src_end + row_offset))
                    
                title_col = mapping_dict.get("C3")
                title = str(row_dict.get(title_col, "")) if title_col and title_col != "(Use Template Default)" else str(row_dict.get("Title of the object", ""))
                
                acc_col = mapping_dict.get("C16")
                accession_no = str(row_dict.get(acc_col, "")) if acc_col and acc_col != "(Use Template Default)" else ""
                if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
                    accession_no = str(row_dict.get("Field Accession No.", ""))
                if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
                    accession_no = str(row_dict.get("NMMA No.", ""))
                
                rows_status.append({
                    "index": row_idx,
                    "success": True,
                    "title": title if title and title.lower() != "nan" else "Unknown Title",
                    "resolved_acc": accession_no if accession_no and accession_no.lower() != "nan" else "None",
                    "missing_crucial": missing_fields,
                    "photo_status": photo_status_msg,
                    "save_status": "Success",
                    "error": None
                })

        # ---------------------------------------------
        # Layout: Horizontally Stacked in a Single Sheet
        # ---------------------------------------------
        elif group_layout == "horizontal":
            ws = wb.active
            ws.title = "NMMA Forms"
            
            if ws.views.sheetView:
                ws.views.sheetView[0].showGridLines = True
            
            src_start_col, src_end_col = 1, 4  # columns A-D
            block_width = src_end_col - src_start_col + 1
            spacing = 1  # 1 blank column between cards
            
            for i, row_dict in enumerate(rows_list):
                col_offset = i * (block_width + spacing)
                row_idx = start_idx + i
                
                if i > 0:
                    copy_block_horizontal(ws, src_start_col, src_end_col, src_start_col + col_offset)
                
                missing_fields = audit_record(row_dict, mapping_dict)
                record = pd.Series(row_dict)
                photo_files = find_photo_files(record, row_idx, photos_dir, mapping_dict)
                photo_path = get_record_photo(record, row_idx, photos_dir, mapping_dict, output_folder)
                
                if photo_path:
                    if len(photo_files) > 1:
                        photo_status_msg = f"Photo matched & merged ({len(photo_files)} files: {', '.join([os.path.basename(p) for p in photo_files])})"
                    else:
                        photo_status_msg = f"Photo matched: {os.path.basename(photo_path)}"
                else:
                    photo_status_msg = "No matching photo found"
                
                mapper = TemplateMapper(wb)
                mapper.sheet = ws
                mapper.fill_record(record, mapping_dict, col_offset=col_offset, photo_path=photo_path, photo_width=photo_width, photo_height=photo_height)
                
                # Column page break
                if i < len(rows_list) - 1:
                    from openpyxl.worksheet.pagebreak import Break
                    ws.col_breaks.append(Break(id=src_end_col + col_offset))
                    
                title_col = mapping_dict.get("C3")
                title = str(row_dict.get(title_col, "")) if title_col and title_col != "(Use Template Default)" else str(row_dict.get("Title of the object", ""))
                
                acc_col = mapping_dict.get("C16")
                accession_no = str(row_dict.get(acc_col, "")) if acc_col and acc_col != "(Use Template Default)" else ""
                if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
                    accession_no = str(row_dict.get("Field Accession No.", ""))
                if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
                    accession_no = str(row_dict.get("NMMA No.", ""))
                
                rows_status.append({
                    "index": row_idx,
                    "success": True,
                    "title": title if title and title.lower() != "nan" else "Unknown Title",
                    "resolved_acc": accession_no if accession_no and accession_no.lower() != "nan" else "None",
                    "missing_crucial": missing_fields,
                    "photo_status": photo_status_msg,
                    "save_status": "Success",
                    "error": None
                })
                    
        # ---------------------------------------------
        # Layout: Separate Sheets (Tabs)
        # ---------------------------------------------
        else:
            template_sheet = wb.active
            
            for i, row_dict in enumerate(rows_list):
                row_idx = start_idx + i
                ws_copy = wb.copy_worksheet(template_sheet)
                
                missing_fields = audit_record(row_dict, mapping_dict)
                record = pd.Series(row_dict)
                photo_files = find_photo_files(record, row_idx, photos_dir, mapping_dict)
                photo_path = get_record_photo(record, row_idx, photos_dir, mapping_dict, output_folder)
                
                if photo_path:
                    if len(photo_files) > 1:
                        photo_status_msg = f"Photo matched & merged ({len(photo_files)} files: {', '.join([os.path.basename(p) for p in photo_files])})"
                    else:
                        photo_status_msg = f"Photo matched: {os.path.basename(photo_path)}"
                else:
                    photo_status_msg = "No matching photo found"
                
                title_col = mapping_dict.get("C3")
                title = str(row_dict.get(title_col, "")) if title_col and title_col != "(Use Template Default)" else str(row_dict.get("Title of the object", ""))
                
                acc_col = mapping_dict.get("C16")
                accession_no = str(row_dict.get(acc_col, "")) if acc_col and acc_col != "(Use Template Default)" else ""
                if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
                    accession_no = str(row_dict.get("Field Accession No.", ""))
                if not accession_no or accession_no.lower() == "nan" or not accession_no.strip():
                    accession_no = str(row_dict.get("NMMA No.", ""))
                
                def sanitize_sheet_name(val):
                    val = str(val).strip()
                    if not val or val.lower() == "nan":
                        return ""
                    for char in r'\/?*::[]':
                        val = val.replace(char, "_")
                    return val

                san_acc = sanitize_sheet_name(accession_no)
                san_title = sanitize_sheet_name(title)
                
                sheet_name_parts = [f"{row_idx:05d}"]
                if san_acc:
                    sheet_name_parts.append(san_acc)
                elif san_title:
                    sheet_name_parts.append(san_title)
                    
                sheet_name = "_".join(sheet_name_parts)[:31]
                
                base_sheet_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    suffix = f"_{counter}"
                    sheet_name = base_sheet_name[:31 - len(suffix)] + suffix
                    counter += 1
                    
                ws_copy.title = sheet_name
                
                mapper = TemplateMapper(wb)
                mapper.sheet = ws_copy
                mapper.fill_record(record, mapping_dict, photo_path=photo_path, photo_width=photo_width, photo_height=photo_height)
                
                rows_status.append({
                    "index": row_idx,
                    "success": True,
                    "title": title if title and title.lower() != "nan" else "Unknown Title",
                    "resolved_acc": accession_no if accession_no and accession_no.lower() != "nan" else "None",
                    "missing_crucial": missing_fields,
                    "photo_status": photo_status_msg,
                    "save_status": "Success",
                    "error": None
                })
                
            wb.remove(template_sheet)
        
        filename = f"NMMA_Forms_{start_idx:05d}_to_{end_idx:05d}.xlsx"
        output_file = os.path.join(output_folder, filename)
        
        save_status_msg = "Success"
        try:
            wb.save(output_file)
        except PermissionError:
            import time
            base, ext = os.path.splitext(output_file)
            fallback_file = f"{base}_LOCKED_{int(time.time())}{ext}"
            wb.save(fallback_file)
            output_file = fallback_file
            save_status_msg = "Saved as copy (original open in Excel)"
            
        wb.close()
        
        for r_stat in rows_status:
            r_stat["save_status"] = save_status_msg
            r_stat["output_file"] = output_file
        
        return {
            "chunk_index": chunk_index,
            "success": True,
            "output_file": output_file,
            "rows_status": rows_status,
            "error": None
        }
    except Exception as e:
        import traceback
        err_msg = f"{str(e)}\n{traceback.format_exc()}"
        failed_rows = []
        for i in range(len(rows_list)):
            row_idx = start_idx + i
            failed_rows.append({
                "index": row_idx,
                "success": False,
                "title": "Unknown Title",
                "resolved_acc": "None",
                "missing_crucial": [],
                "photo_status": "Failed to process photo",
                "save_status": "Failed to save file",
                "error": err_msg
            })
        return {
            "chunk_index": chunk_index,
            "success": False,
            "output_file": None,
            "rows_status": failed_rows,
            "error": err_msg
        }


class ExcelEngine:
    def __init__(self, book_path, template_path, output_folder):
        self.book_path = book_path
        self.template_path = template_path
        self.output_folder = output_folder

    def generate_first_form(self):
        df = pd.read_excel(self.book_path)
        record = df.iloc[0]
        wb = openpyxl.load_workbook(self.template_path)
        mapper = TemplateMapper(wb)
        mapper.fill_record(record, {}, photo_path=None)
        os.makedirs(self.output_folder, exist_ok=True)
        output_file = os.path.join(self.output_folder, "NMMA_0001.xlsx")
        wb.save(output_file)
        wb.close()
        return output_file


class GeneratorWorker(QThread):
    progress_updated = Signal(int, int)  # completed, total
    status_updated = Signal(str)
    finished = Signal(int, int, str)     # success_count, error_count, summary_msg
    error_occurred = Signal(str)

    def __init__(self, book_path, template_path, output_folder, output_format, group_size=1, group_layout="sheets", mapping_dict=None, photos_dir=None, photo_width=220, photo_height=200):
        super().__init__()
        self.book_path = book_path
        self.template_path = template_path
        self.output_folder = output_folder
        self.output_format = output_format
        self.group_size = group_size
        self.group_layout = group_layout
        self.mapping_dict = mapping_dict if mapping_dict is not None else {}
        self.photos_dir = photos_dir
        self.photo_width = photo_width
        self.photo_height = photo_height
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def run(self):
        try:
            self.status_updated.emit("Reading master data file...")
            try:
                df = pd.read_excel(self.book_path)
            except Exception as e:
                self.error_occurred.emit(f"Failed to read master data Excel file:\n{e}")
                return

            total_rows = len(df)
            if total_rows == 0:
                self.error_occurred.emit("The Master data Excel sheet is empty.")
                return

            os.makedirs(self.output_folder, exist_ok=True)

            import multiprocessing
            cpu_count = multiprocessing.cpu_count()
            num_workers = max(1, cpu_count - 1)
            
            success_count = 0
            error_count = 0
            generated_excel_files = []
            results_list = []

            # ---------------------------------------------
            # Grouped Workbooks Mode (group_size > 1)
            # ---------------------------------------------
            if self.group_size > 1:
                chunks = []
                chunk_index = 1
                for start_idx in range(0, total_rows, self.group_size):
                    end_idx = min(start_idx + self.group_size, total_rows)
                    chunk_df = df.iloc[start_idx:end_idx]
                    
                    rows_list = [row.to_dict() for _, row in chunk_df.iterrows()]
                    chunks.append((
                        rows_list,
                        self.template_path,
                        self.output_folder,
                        chunk_index,
                        start_idx + 1,
                        end_idx,
                        self.group_layout,
                        self.mapping_dict,
                        self.photos_dir,
                        self.photo_width,
                        self.photo_height
                    ))
                    chunk_index += 1

                total_chunks = len(chunks)
                self.status_updated.emit(f"Spawning parallel processes for {total_chunks} grouped workbooks...")
                
                pool = multiprocessing.Pool(processes=num_workers)
                results = pool.imap_unordered(process_chunk, chunks)
                
                for res in results:
                    if self._is_cancelled:
                        pool.terminate()
                        pool.join()
                        self.status_updated.emit("Generation cancelled by user.")
                        self.finished.emit(success_count, error_count, "Generation cancelled by user.")
                        return

                    results_list.append(res)
                    if res["success"]:
                        success_count += 1
                        generated_excel_files.append(res["output_file"])
                    else:
                        error_count += 1
                        print(f"Error processing chunk {res['chunk_index']}: {res['error']}")
                        
                    completed_chunks = success_count + error_count
                    self.progress_updated.emit(completed_chunks, total_chunks)
                    self.status_updated.emit(f"Generated {completed_chunks}/{total_chunks} grouped workbooks...")

                pool.close()
                pool.join()

            # ---------------------------------------------
            # Individual Files Mode (group_size <= 1)
            # ---------------------------------------------
            else:
                tasks = []
                for idx, (_, row) in enumerate(df.iterrows(), start=1):
                    row_dict = row.to_dict()
                    tasks.append((row_dict, self.template_path, self.output_folder, idx, self.mapping_dict, self.photos_dir, self.photo_width, self.photo_height))

                self.status_updated.emit(f"Spawning parallel processes for {total_rows} antiquity forms...")
                
                pool = multiprocessing.Pool(processes=num_workers)
                results = pool.imap_unordered(process_single_row, tasks)
                
                for res in results:
                    if self._is_cancelled:
                        pool.terminate()
                        pool.join()
                        self.status_updated.emit("Generation cancelled by user.")
                        self.finished.emit(success_count, error_count, "Generation cancelled by user.")
                        return

                    results_list.append(res)
                    if res["success"]:
                        success_count += 1
                        generated_excel_files.append(res["output_file"])
                    else:
                        error_count += 1
                        print(f"Error processing row index {res['index']}: {res['error']}")

                    processed = success_count + error_count
                    self.progress_updated.emit(processed, total_rows)
                    self.status_updated.emit(f"Generated {processed}/{total_rows} Excel forms...")

                pool.close()
                pool.join()

            # Clean up temporary merged photos directory
            temp_photos_dir = os.path.join(self.output_folder, ".temp_photos")
            if os.path.exists(temp_photos_dir):
                import shutil
                try:
                    shutil.rmtree(temp_photos_dir)
                except Exception:
                    pass

            # ---------------------------------------------
            # Compile Audit Log Records & Write Excel Sheet
            # ---------------------------------------------
            self.status_updated.emit("Compiling data-integrity Audit Report...")
            audit_records = []
            if self.group_size > 1:
                for res in results_list:
                    audit_records.extend(res.get("rows_status", []))
            else:
                for res in results_list:
                    audit_records.append(res)
            
            audit_records = sorted(audit_records, key=lambda x: x["index"])
            
            total_processed = len(audit_records)
            missing_photos_count = sum(1 for r in audit_records if "No matching photo" in r["photo_status"])
            missing_values_count = sum(1 for r in audit_records if len(r["missing_crucial"]) > 0)
            locked_files_count = sum(1 for r in audit_records if "copy" in r["save_status"] or "LOCKED" in r["save_status"])
            failed_count = sum(1 for r in audit_records if not r["success"])
            
            audit_msg = ""
            try:
                audit_rows = []
                for r in audit_records:
                    audit_rows.append({
                        "Row Index": r["index"],
                        "Antiquity Title": r["title"],
                        "Resolved Accession No.": r["resolved_acc"],
                        "Crucial Fields Missing": ", ".join(r["missing_crucial"]) if r["missing_crucial"] else "None (OK)",
                        "Photo Status": r["photo_status"],
                        "Save Status": r["save_status"],
                        "Status": "Success" if r["success"] else "Failed",
                        "Error Details": r["error"] if r["error"] else ""
                    })
                df_audit = pd.DataFrame(audit_rows)
                audit_report_path = os.path.join(self.output_folder, "Generation_Audit_Report.xlsx")
                df_audit.to_excel(audit_report_path, index=False)
                audit_msg = f"Audit Report saved to:\n{os.path.basename(audit_report_path)}"
            except Exception as e_audit:
                audit_msg = f"Failed to save Audit Report: {e_audit}"

            # PDF export phase
            if self.output_format in ["PDF", "Excel + PDF"] and not self._is_cancelled:
                self.status_updated.emit("Initializing Microsoft Excel for PDF conversion...")
                self.progress_updated.emit(0, len(generated_excel_files))

                from pdf_exporter import ExcelPDFConverter
                converter = ExcelPDFConverter()
                try:
                    converter.start()
                except Exception as e:
                    self.error_occurred.emit(str(e))
                    return

                pdf_success = 0
                pdf_error = 0
                total_pdfs = len(generated_excel_files)

                for idx, xlsx_path in enumerate(generated_excel_files, start=1):
                    if self._is_cancelled:
                        break

                    self.status_updated.emit(f"Converting to PDF ({idx}/{total_pdfs}): {os.path.basename(xlsx_path)}")
                    
                    try:
                        pdf_filename = os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
                        pdf_path = os.path.join(self.output_folder, pdf_filename)
                        
                        converter.convert(xlsx_path, pdf_path)
                        pdf_success += 1
                        
                        if self.output_format == "PDF":
                            try:
                                os.remove(xlsx_path)
                            except Exception:
                                pass
                    except Exception as e:
                        pdf_error += 1
                        print(f"PDF conversion failed for {xlsx_path}: {e}")

                    self.progress_updated.emit(idx, total_pdfs)

                converter.stop()

                if self._is_cancelled:
                    self.status_updated.emit("Cancelled during PDF export.")
                    self.finished.emit(success_count, error_count, "PDF conversion cancelled by user.")
                    return

                summary = (
                    f"Completed successfully!\n\n"
                    f"Antiquities Processed: {total_processed} (Failed: {failed_count})\n"
                    f"Excel Files Saved: {success_count} (Failed: {error_count})\n"
                    f"PDF Files Generated: {pdf_success} (Failed: {pdf_error})\n\n"
                    f"Data Integrity Summary:\n"
                    f"  - Missing Photos: {missing_photos_count}\n"
                    f"  - Missing Crucial Values: {missing_values_count}\n"
                    f"  - Locked Files Recovered: {locked_files_count}\n\n"
                    f"{audit_msg}"
                )
                self.finished.emit(success_count, error_count, summary)

            else:
                summary = (
                    f"Completed successfully!\n\n"
                    f"Antiquities Processed: {total_processed} (Failed: {failed_count})\n"
                    f"Excel Files Saved: {success_count} (Failed: {error_count})\n\n"
                    f"Data Integrity Summary:\n"
                    f"  - Missing Photos: {missing_photos_count}\n"
                    f"  - Missing Crucial Values: {missing_values_count}\n"
                    f"  - Locked Files Recovered: {locked_files_count}\n\n"
                    f"{audit_msg}"
                )
                self.finished.emit(success_count, error_count, summary)

        except Exception as e:
            import traceback
            self.error_occurred.emit(f"Unexpected worker error:\n{e}\n{traceback.format_exc()}")